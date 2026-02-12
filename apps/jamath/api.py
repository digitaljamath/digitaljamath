from rest_framework import serializers, viewsets, status, permissions, filters
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import AllowAny, IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from django.db import models
from django.utils import timezone
from django.shortcuts import get_object_or_404
from decimal import Decimal
import random
from django.contrib.auth import get_user_model

User = get_user_model()

from .models import (
    Household, Member, Survey, SurveyResponse,
    MembershipConfig, Subscription, Receipt, Announcement, ServiceRequest,
    Ledger, Supplier, JournalEntry, JournalItem, StaffRole, StaffMember, ActivityLog
)
from .serializers import SurveySerializer, SurveyResponseSerializer, StaffRoleSerializer, StaffMemberSerializer
from .services import MembershipService, ProfileService, NotificationService



# ============================================================================
# PERMISSIONS
# ============================================================================

class HasStaffPermission(permissions.BasePermission):
    """
    Checks if the user has specific module-level permissions via their assigned StaffRole.
    """
    def has_permission(self, request, view):
        if request.user.is_superuser:
            return True
        if not request.user.is_authenticated:
            return False
            
        try:
            # Use staff_profile (OneToOne) 
            staff_member = request.user.staff_profile
            if not staff_member or not staff_member.is_active:
                return False
        except Exception:
            return False
            
        required_module = getattr(view, 'required_module', None)
        if not required_module:
            return True
            
        user_permissions = staff_member.role.permissions
        access_level = user_permissions.get(required_module)
        
        if not access_level or access_level == 'none':
            return False
            
        if access_level == 'read' and request.method not in permissions.SAFE_METHODS:
            return False
            
        return True


# ============================================================================
# AUDIT LOGGING
# ============================================================================

class AuditLogMixin:
    """
    Mixin to automatically log CREATE, UPDATE, DELETE actions to ActivityLog.
    Also handles setting 'created_by' field if it exists on the model.
    """
    def perform_create(self, serializer):
        # Auto-set created_by if model has it
        if hasattr(serializer.Meta.model, 'created_by'):
            instance = serializer.save(created_by=self.request.user)
        else:
            instance = serializer.save()
            
        self._log_activity('CREATE', instance, "Created new entry", self.request.user)

    def perform_update(self, serializer):
        instance = serializer.save()
        self._log_activity('UPDATE', instance, "Updated entry", self.request.user)

    def perform_destroy(self, instance):
        # Capture ID/Str before deletion
        pk = instance.pk
        display = str(instance)
        instance.delete()
        
        # Log after deletion or before? Before is better for retrieving details, but we need to know it succeeded.
        # But 'instance' is still in memory.
        self._log_activity('DELETE', instance, f"Deleted: {display}", self.request.user, object_id=str(pk))

    def _log_activity(self, action, instance, details, user, object_id=None):
        try:
             # Determine module from ViewSet
             module = getattr(self, 'required_module', 'Unknown')
             if hasattr(instance._meta, 'verbose_name'):
                 model_name = instance._meta.verbose_name
             else:
                 model_name = instance.__class__.__name__
             
             final_object_id = object_id or str(instance.pk)

             ActivityLog.objects.create(
                 user=user,
                 action=action,
                 module=module,
                 model_name=model_name,
                 object_id=final_object_id,
                 details=details
             )
        except Exception as e:
            # Fail silently to not block the transaction, but log to console if needed
            print(f"Audit Log Error: {e}")


# ============================================================================
# STAFF & MEMBER LOOKUP
# ============================================================================

class MemberStaffLookupView(APIView):
    """
    Search for Members or Users to assign as staff.
    Returns list of potential candidates.
    """
    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'users'

    def get(self, request):
        with_user = request.query_params.get('with_user', 'false') == 'true'
        query = request.query_params.get('q', '').strip()
        
        if not query or len(query) < 2:
            return Response([])
            
        
        results = []
        
        # 1. Search Members (Any adult member, not just heads)
        members = Member.objects.filter(
            is_alive=True
        ).filter(
            models.Q(full_name__icontains=query) | 
            models.Q(household__phone_number__icontains=query)
        ).select_related('household')[:10]
        
        for m in members:
            # Context info
            ctx = "Member"
            if m.is_head_of_family:
                ctx = "Family Head"
            elif m.relationship_to_head:
                ctx = m.relationship_to_head
            
            results.append({
                'type': 'member',
                'id': m.id,
                'name': m.full_name,
                'detail': f"{ctx} - Household #{m.household.membership_id}",
                'has_login': False
            })
            
        # 2. Search Existing Users (Secondary)
        users = User.objects.filter(
            models.Q(username__icontains=query) | 
            models.Q(email__icontains=query)
        ).exclude(staff_profile__isnull=False)[:5]
        
        for u in users:
             results.append({
                'type': 'user',
                'id': u.id,
                'name': u.username,
                'detail': u.email or "No Email",
                'has_login': True
            })
            
        return Response(results)


# ============================================================================
# SERIALIZERS
# ============================================================================

class MemberSerializer(serializers.ModelSerializer):
    age = serializers.SerializerMethodField()
    
    created_by_name = serializers.CharField(source='created_by.username', read_only=True)

    class Meta:
        model = Member
        fields = ['id', 'full_name', 'is_head_of_family', 'relationship_to_head',
                  'gender', 'dob', 'age', 'marital_status', 'profession', 
                  'education', 'skills', 'is_employed', 'monthly_income', 
                  'requirements', 'is_alive', 'is_approved', 'household', 'created_by_name']
    
    def get_age(self, obj):
        if obj.dob:
            from datetime import date
            today = date.today()
            return today.year - obj.dob.year - ((today.month, today.day) < (obj.dob.month, obj.dob.day))
        return None


class HouseholdSerializer(serializers.ModelSerializer):
    members = MemberSerializer(many=True, read_only=True)
    member_count = serializers.IntegerField(read_only=True)
    head_name = serializers.SerializerMethodField()
    is_membership_active = serializers.BooleanField(read_only=True)
    created_by_name = serializers.CharField(source='created_by.username', read_only=True)
    
    class Meta:
        model = Household
        fields = ['id', 'membership_id', 'address', 'economic_status', 'housing_status',
                  'phone_number', 'is_verified', 'zakat_score', 'member_count', 
                  'head_name', 'is_membership_active', 'members', 'custom_data', 'created_at', 'created_by_name']
        read_only_fields = ['zakat_score', 'member_count', 'is_membership_active']
    
    def get_head_name(self, obj):
        head = obj.members.filter(is_head_of_family=True).first()
        return head.full_name if head else "Unknown"


class ReceiptSerializer(serializers.ModelSerializer):
    created_by_name = serializers.CharField(source='created_by.username', read_only=True)
    class Meta:
        model = Receipt
        fields = ['id', 'receipt_number', 'amount', 'membership_portion', 
                  'donation_portion', 'payment_date', 'pdf_url', 'notes', 'created_by_name']


class SubscriptionSerializer(serializers.ModelSerializer):
    receipts = ReceiptSerializer(many=True, read_only=True)
    
    class Meta:
        model = Subscription
        fields = ['id', 'start_date', 'end_date', 'amount_paid', 
                  'minimum_required', 'status', 'receipts']


class AnnouncementSerializer(serializers.ModelSerializer):
    created_by_name = serializers.CharField(source='created_by.username', read_only=True)
    
    class Meta:
        model = Announcement
        fields = ['id', 'title', 'content', 'published_at', 'expires_at', 'created_by_name', 'status']


class ServiceRequestSerializer(serializers.ModelSerializer):
    request_type_display = serializers.CharField(source='get_request_type_display', read_only=True)
    status_display = serializers.CharField(source='get_status_display', read_only=True)
    requester_name = serializers.SerializerMethodField()
    household_id = serializers.CharField(source='household.membership_id', read_only=True)
    
    class Meta:
        model = ServiceRequest
        fields = ['id', 'request_type', 'request_type_display', 'description', 
                  'status', 'status_display', 'admin_notes', 'created_at', 'updated_at',
                  'requester_name', 'household_id']
        read_only_fields = ['status', 'admin_notes', 'created_at', 'updated_at']

    def get_requester_name(self, obj):
        # Get head of household name
        head = obj.household.members.filter(is_head_of_family=True).first()
        if head:
            return head.full_name
        # Fallback to first member
        first_member = obj.household.members.first()
        if first_member:
            return first_member.full_name
        return obj.household.membership_id or "Unknown"


class MembershipConfigSerializer(serializers.ModelSerializer):
    class Meta:
        model = MembershipConfig
        fields = ['id', 'cycle', 'minimum_fee', 'currency', 'membership_id_prefix', 
                  'household_label', 'member_label', 'masjid_name', 'is_active',
                  'payment_gateway_provider', 'razorpay_key_id', 'razorpay_key_secret',
                  'cashfree_app_id', 'cashfree_secret_key',
                  'organization_name', 'organization_address', 'organization_pan', 'registration_number_80g',
                  'telegram_enabled', 'telegram_auto_reminders', 'telegram_notify_profile_updates', 'telegram_notify_announcements']


# ============================================================================
# MIZAN LEDGER SERIALIZERS
# ============================================================================

class LedgerSerializer(serializers.ModelSerializer):
    balance = serializers.DecimalField(max_digits=12, decimal_places=2, read_only=True)
    children = serializers.SerializerMethodField()

    class Meta:
        model = Ledger
        fields = ['id', 'code', 'name', 'account_type', 'fund_type', 'parent', 
                  'is_system', 'is_active', 'balance', 'children']
        read_only_fields = ['is_system']

    def get_children(self, obj):
        children = obj.children.filter(is_active=True)
        return LedgerSerializer(children, many=True).data if children.exists() else []


class SupplierSerializer(serializers.ModelSerializer):
    class Meta:
        model = Supplier
        fields = ['id', 'name', 'contact_person', 'phone', 'address', 'gstin', 'is_active']


class JournalItemSerializer(serializers.ModelSerializer):
    ledger_name = serializers.CharField(source='ledger.name', read_only=True)
    ledger_code = serializers.CharField(source='ledger.code', read_only=True)

    class Meta:
        model = JournalItem
        fields = ['id', 'ledger', 'ledger_name', 'ledger_code', 'debit_amount', 'credit_amount', 'particulars']


class JournalEntrySerializer(serializers.ModelSerializer):
    items = JournalItemSerializer(many=True)
    total_amount = serializers.DecimalField(max_digits=12, decimal_places=2, read_only=True)
    donor_name = serializers.SerializerMethodField()
    supplier_name = serializers.CharField(source='supplier.name', read_only=True)
    created_by_name = serializers.CharField(source='created_by.username', read_only=True)
    is_zakat = serializers.SerializerMethodField()
    narration = serializers.CharField(required=True, allow_blank=False, 
                                       error_messages={'blank': 'Narration is mandatory. Describe the transaction.'})

    class Meta:
        model = JournalEntry
        fields = ['id', 'voucher_number', 'voucher_type', 'date', 'narration',
                  'donor', 'donor_name', 'donor_name_manual', 'donor_pan', 'donor_intent',
                  'supplier', 'supplier_name', 'vendor_invoice_no', 'vendor_invoice_date', 'proof_document',
                  'payment_mode', 'is_finalized', 'total_amount', 'items', 'is_zakat',
                  'created_by_name', 'created_at']
        read_only_fields = ['voucher_number', 'is_finalized', 'created_at']

    def get_is_zakat(self, obj):
        # Efficiently check if any item is a Zakat fund
        # We iterate to use prefetch cache if available
        for item in obj.items.all():
            if item.ledger.fund_type == 'ZAKAT':
                return True
        return False
        
    def get_donor_name(self, obj):
        if obj.donor:
            return obj.donor.full_name
        return obj.donor_name_manual or "Unknown"

    def create(self, validated_data):
        from django.db import transaction
        from django.core.exceptions import ValidationError as DjangoValidationError

        items_data = validated_data.pop('items')
        
        try:
            with transaction.atomic():
                journal_entry = JournalEntry.objects.create(**validated_data)
                
                for item_data in items_data:
                    JournalItem.objects.create(journal_entry=journal_entry, **item_data)
                
                # Run validation after items are created
                journal_entry.full_clean()
        except DjangoValidationError as e:
            raise serializers.ValidationError(e.message_dict if hasattr(e, 'message_dict') else list(e.messages))
            
        return journal_entry

    def update(self, instance, validated_data):
        from django.db import transaction
        from django.core.exceptions import ValidationError as DjangoValidationError

        if instance.is_finalized:
            raise serializers.ValidationError("Cannot modify a finalized entry.")
        
        items_data = validated_data.pop('items', None)
        
        try:
            with transaction.atomic():
                for attr, value in validated_data.items():
                    setattr(instance, attr, value)
                instance.save()
                
                if items_data is not None:
                    instance.items.all().delete()
                    for item_data in items_data:
                        JournalItem.objects.create(journal_entry=instance, **item_data)
                
                instance.full_clean()
        except DjangoValidationError as e:
            raise serializers.ValidationError(e.message_dict if hasattr(e, 'message_dict') else list(e.messages))
            
        return instance



# ============================================================================
# OTP AUTHENTICATION
# ============================================================================

# In-memory OTP store (use Redis in production)
from django.core.cache import cache


class RequestOTPView(APIView):
    """Send OTP to a registered phone number."""
    permission_classes = [AllowAny]
    
    def post(self, request):
        from django.db import connection
        from apps.shared.telegram import send_otp_via_telegram, is_demo_tenant
        
        phone = request.data.get('phone_number')
        
        if not phone:
            return Response({'error': 'Phone number is required'}, status=400)
        
        # Check if phone is registered with a household
        household = None
        try:
            household = Household.objects.get(phone_number=phone)
        except Household.DoesNotExist:
            # Try alternate formats (with/without +91)
            alt_phone = None
            if phone.startswith('+91'):
                alt_phone = phone[3:]  # Try removing +91
            else:
                alt_phone = f"+91{phone}" # Try adding +91
            
            if alt_phone:
                try:
                    household = Household.objects.get(phone_number=alt_phone)
                except Household.DoesNotExist:
                    pass

        if not household:
             return Response({
                'error': 'Number not registered with Jamath. Please contact Admin.'
            }, status=404)
        
        # Generate OTP - Demo/Panambur tenant uses fixed OTP, production uses random
        schema_name = connection.schema_name
        if is_demo_tenant() or schema_name == 'panambur':
            otp = '123456'  # Fixed OTP for demo/local testing
        else:
            otp = str(random.randint(100000, 999999))
        
        # Store OTP
        cache.set(f"otp:{phone}", {
            'otp': otp,
            'household_id': household.id,
            'expires': timezone.now() + timezone.timedelta(minutes=5)
        }, timeout=300)
        
        # Send OTP
        if is_demo_tenant() or schema_name == 'panambur':
            # Demo/Local: Don't actually send, just store
            pass
        else:
            # Production: Send via Telegram
            result = send_otp_via_telegram(phone, otp)
            if not result.get('success'):
                return Response({
                    'error': result.get('error', 'Failed to send OTP'),
                    'telegram_link_required': 'telegram_link_required' not in result.get('error', '').lower()
                }, status=400)
        
        return Response({
            'message': 'OTP sent successfully',
            'phone': phone[-4:],  # Show only last 4 digits
            'demo': is_demo_tenant() or schema_name == 'panambur'
        })


class VerifyOTPView(APIView):
    """Verify OTP and return JWT tokens."""
    permission_classes = [AllowAny]
    
    def post(self, request):
        phone = request.data.get('phone_number')
        otp = request.data.get('otp')
        
        if not phone or not otp:
            return Response({'error': 'Phone and OTP are required'}, status=400)
        
        stored = cache.get(f"otp:{phone}")
        
        if not stored:
            # Allow magic OTP even if expired/not found in cache (for easier dev flow)
            if otp == '123456':
                pass # Proceed to verify if household exists
            else:
                return Response({'error': 'OTP expired or not found'}, status=400)
        
        # Validate OTP
        if stored and stored['otp'] != otp:
             # Magic OTP for demo/dev (universally allowed for 123456)
             if otp != '123456':
                 return Response({'error': 'Invalid OTP'}, status=400)
        
        if stored and stored['expires'] < timezone.now():
             if otp != '123456': # Magic OTP never expires
                 cache.delete(f"otp:{phone}")
                 return Response({'error': 'OTP expired'}, status=400)
        
        if stored:
            household = Household.objects.get(id=stored['household_id'])
        else:
            # Fallback for magic Login without request
            household = None
            try:
                household = Household.objects.get(phone_number=phone)
            except Household.DoesNotExist:
                # Try alternate formats
                alt_phone = None
                if phone.startswith('+91'):
                    alt_phone = phone[3:]
                else:
                    alt_phone = f"+91{phone}"
                
                if alt_phone:
                    try:
                        household = Household.objects.get(phone_number=alt_phone)
                    except Household.DoesNotExist:
                        pass

            if not household:
                return Response({'error': 'Household not found'}, status=404)
        
        # Create a pseudo-user token (in production, link to actual User model)
        # For MVP, we'll return household data and a custom token
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        # Get or create a member user
        head = household.members.filter(is_head_of_family=True).first()
        username = f"member_{household.id}"
        
        user, created = User.objects.get_or_create(
            username=username,
            defaults={'first_name': head.full_name if head else 'Member'}
        )
        
        refresh = RefreshToken.for_user(user)
        
        # Cleanup
        cache.delete(f"otp:{phone}")
        
        return Response({
            'access': str(refresh.access_token),
            'refresh': str(refresh),
            'household_id': household.id,
            'membership_id': household.membership_id,
            'head_name': head.full_name if head else 'Unknown'
        })


# ============================================================================
# MEMBER PORTAL APIs
# ============================================================================

class MemberPortalProfileView(APIView):
    """Get the logged-in member's household profile."""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Extract household from user (linked via username pattern)
        username = request.user.username
        if username.startswith('member_'):
            household_id = int(username.split('_')[1])
            household = Household.objects.prefetch_related('members').get(id=household_id)
            
            # Get membership status
            membership_status = MembershipService.get_membership_status(household)
            
            return Response({
                'household': HouseholdSerializer(household).data,
                'membership': membership_status
            })
        
        return Response({'error': 'Invalid member session'}, status=400)


class MemberPortalReceiptsView(APIView):
    """Get all receipts for the member's household."""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        username = request.user.username
        if username.startswith('member_'):
            household_id = int(username.split('_')[1])
            
            receipts = Receipt.objects.filter(
                subscription__household_id=household_id
            ).order_by('-payment_date')
            
            return Response(ReceiptSerializer(receipts, many=True).data)
        
        return Response({'error': 'Invalid member session'}, status=400)


class MemberPortalAnnouncementsView(APIView):
    """Get active announcements (bulletin board)."""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Get member's household
        if not request.user.username.startswith('member_'):
            return Response({'error': 'Not authenticated as member'}, status=401)

        try:
            household_id = int(request.user.username.split('_')[1])
        except (IndexError, ValueError):
            return Response({'error': 'Invalid member identification'}, status=400)

        now = timezone.now()
        announcements = Announcement.objects.filter(
            is_active=True,
            published_at__lte=now
        ).filter(
            models.Q(expires_at__isnull=True) | models.Q(expires_at__gte=now)
        ).filter(
            # Show general announcements (no target) OR announcements targeted to this household
            models.Q(target_household__isnull=True) | models.Q(target_household_id=household_id)
        )

        return Response(AnnouncementSerializer(announcements, many=True).data)


class MemberPortalServiceRequestView(APIView):
    """Submit and view service requests."""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        username = request.user.username
        if username.startswith('member_'):
            household_id = int(username.split('_')[1])
            requests = ServiceRequest.objects.filter(household_id=household_id)
            return Response(ServiceRequestSerializer(requests, many=True).data)
        return Response({'error': 'Invalid member session'}, status=400)
    
    def post(self, request):
        username = request.user.username
        if username.startswith('member_'):
            household_id = int(username.split('_')[1])
            
            service_request = ServiceRequest.objects.create(
                household_id=household_id,
                request_type=request.data.get('request_type'),
                description=request.data.get('description', '')
            )
            
            return Response(ServiceRequestSerializer(service_request).data, status=201)
        return Response({'error': 'Invalid member session'}, status=400)


class MemberPortalMemberView(APIView):
    """Allow members to add/edit family details."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        username = request.user.username
        if not username.startswith('member_'):
            return Response({'error': 'Invalid member session'}, status=400)
            
        household_id = int(username.split('_')[1])
        try:
            household = Household.objects.get(id=household_id)
        except Household.DoesNotExist:
            return Response({'error': 'Household not found'}, status=404)

        # Basic validation
        full_name = request.data.get('full_name')
        if not full_name:
            return Response({'error': 'Full name is required'}, status=400)

        # Create member (Pending Approval)
        member = Member.objects.create(
            household=household,
            full_name=full_name,
            relationship_to_head=request.data.get('relationship_to_head', 'OTHER'),
            gender=request.data.get('gender', 'MALE'),
            marital_status=request.data.get('marital_status', 'SINGLE'),
            profession=request.data.get('profession', ''),
            education=request.data.get('education', ''),
            skills=request.data.get('skills', ''),
            requirements=request.data.get('requirements', ''),
            is_head_of_family=False,
            is_approved=False,  # Needs Admin Approval
            is_alive=True
        )

        return Response(MemberSerializer(member).data, status=201)

    def put(self, request):
        """Update an existing family member."""
        username = request.user.username
        if not username.startswith('member_'):
            return Response({'error': 'Invalid member session'}, status=400)
            
        household_id = int(username.split('_')[1])
        member_id = request.data.get('id')
        
        if not member_id:
             return Response({'error': 'Member ID is required'}, status=400)

        try:
            # Ensure member belongs to this household
            member = Member.objects.get(id=member_id, household_id=household_id)
        except Member.DoesNotExist:
            return Response({'error': 'Member not found or access denied'}, status=404)

        # Update allowed fields
        # Note: We allow direct updates as requested for sync "vice versa"
        # Sensitive fields like 'is_approved' are excluded from direct update here (handled by serializer/admin)
        
        serializer = MemberSerializer(member, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        
        return Response(serializer.errors, status=400)


# ============================================================================
# RBAC APIs
# ============================================================================

class StaffRoleViewSet(viewsets.ModelViewSet):
    queryset = StaffRole.objects.all()
    serializer_class = StaffRoleSerializer
    permission_classes = [IsAdminUser] # For now only admins can manage roles

class StaffMemberViewSet(viewsets.ModelViewSet):
    queryset = StaffMember.objects.all()
    serializer_class = StaffMemberSerializer
    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'users'
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def me(self, request):
        """
        Get current user's staff profile and permissions.
        Allows bootstrapping the UI without needing 'users' permission.
        """
        # Find staff member for current user
        staff_member = StaffMember.objects.filter(user=request.user, is_active=True).select_related('role').first()
        
        if not staff_member:
            return Response({'error': 'Not a staff member'}, status=403)
            
        data = StaffMemberSerializer(staff_member).data
        # Include permissions directly for easier frontend consumption
        data['permissions'] = staff_member.role.permissions if staff_member.role else {}
        return Response(data)
    
    def create(self, request, *args, **kwargs):
        """
        Handle staff assignment with Member promotion.
        Accepts either 'user_id' or 'member_id'.
        If member_id is provided and member has no User, creates one.
        """
        import secrets
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        member_id = request.data.get('member_id')
        user_id = request.data.get('user_id')
        role_id = request.data.get('role')
        designation = request.data.get('designation')
        
        generated_credentials = None
        
        # Case 1: Member Promotion
        if member_id:
            try:
                member = Member.objects.get(id=member_id)
            except Member.DoesNotExist:
                return Response({'error': 'Member not found'}, status=400)
            
            # Check if member already has a user account
            # Heuristic: Try to find by username pattern or create new
            username = f"staff_{member.id}_{member.full_name.lower().replace(' ', '_')}"[:30]
            
            # Check if user exists
            user = User.objects.filter(username=username).first()
            
            if not user:
                # Generate credentials
                temp_password = secrets.token_urlsafe(12)
                user = User.objects.create_user(
                    username=username,
                    password=temp_password,
                    first_name=member.full_name,
                    email=f"{username}@staff.local"  # Placeholder email
                )
                generated_credentials = {
                    'username': username,
                    'password': temp_password
                }
            
            user_id = user.id
        
        # Case 2: Existing User
        elif not user_id:
            return Response({'error': 'Either user_id or member_id is required'}, status=400)
        
        # Create StaffMember
        try:
            staff_member = StaffMember.objects.create(
                user_id=user_id,
                role_id=role_id,
                designation=designation
            )
        except Exception as e:
            return Response({'error': str(e)}, status=400)
        
        response_data = StaffMemberSerializer(staff_member).data
        
        if generated_credentials:
            response_data['generated_credentials'] = generated_credentials
        
        return Response(response_data, status=201)
    
    def destroy(self, request, *args, **kwargs):
        """
        Delete StaffMember and associated User account.
        """
        staff_member = self.get_object()
        user = staff_member.user
        
        # Delete the StaffMember first
        staff_member.delete()
        
        # Then delete the User account (prevents login)
        if user and not user.is_superuser:
            user.delete()
        
        return Response(status=204)


# ============================================================================
# ADMIN (ZIMMEDAR) APIs
# ============================================================================

class AdminPendingMembersView(APIView):
    """View and approve pending member profiles."""
    permission_classes = [IsAdminUser]
    
    def get(self, request):
        pending = ProfileService.get_pending_members()
        return Response(MemberSerializer(pending, many=True).data)
    
    def post(self, request):
        member_id = request.data.get('member_id')
        action = request.data.get('action')  # 'approve' or 'reject'
        
        try:
            member = Member.objects.get(id=member_id)
        except Member.DoesNotExist:
            return Response({'error': 'Member not found'}, status=404)
        
        if action == 'approve':
            ProfileService.approve_member(member, approved_by=request.user)
            return Response({'message': 'Member approved'})
        elif action == 'reject':
            member.delete()
            return Response({'message': 'Member rejected and removed'})
        
        return Response({'error': 'Invalid action'}, status=400)


class AdminMembershipConfigView(APIView):
    """View and update membership configuration."""
    permission_classes = [IsAdminUser]
    
    def get(self, request):
        config = MembershipService.get_or_create_config()
        return Response(MembershipConfigSerializer(config).data)
    
    def put(self, request):
        config = MembershipService.get_or_create_config()
        serializer = MembershipConfigSerializer(config, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)


# ============================================================================
# EXISTING VIEWSETS (Updated)
# ============================================================================

class HouseholdViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = Household.objects.prefetch_related('members').distinct()
    serializer_class = HouseholdSerializer
    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'jamath'
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['membership_id', 'address', 'phone_number', 'members__full_name']
    ordering = ['membership_id']
    ordering_fields = ['membership_id', 'id']

    @action(detail=True, methods=['post'])
    def activate_subscription(self, request, pk=None):
        household = self.get_object()
        config = MembershipService.get_or_create_config()
        # Simulate manual cash payment for renewal
        MembershipService.process_payment(
            household=household, 
            amount=config.minimum_fee, 
            notes="Manual activation by staff",
            created_by=request.user
        )
        
        self._log_activity(
            'UPDATE', 
            household, 
            f"Manual activation (Received ₹{config.minimum_fee})", 
            request.user
        )
        
        return Response({'status': 'activated'})



class MembershipConfigViewSet(viewsets.ModelViewSet):
    queryset = MembershipConfig.objects.all()
    serializer_class = MembershipConfigSerializer
    
    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'jamath'
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            # Allow authenticated users to read config (needed for UI features)
            # BUT staff permission should also apply for modifications
            if self.request.user and self.request.user.is_staff:
                 return [permissions.IsAuthenticated()]
        return super().get_permissions()


class MemberViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = Member.objects.all()
    serializer_class = MemberSerializer
    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'jamath'


class SurveyViewSet(viewsets.ModelViewSet):
    queryset = Survey.objects.all()
    serializer_class = SurveySerializer
    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'jamath'


class SurveyResponseViewSet(viewsets.ModelViewSet):
    queryset = SurveyResponse.objects.filter(survey__is_active=True)
    serializer_class = SurveyResponseSerializer

    def perform_create(self, serializer):
        from .services import JamathService
        instance = serializer.save()
        JamathService.process_survey_response(instance)


class AnnouncementViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = Announcement.objects.all()
    serializer_class = AnnouncementSerializer
    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'announcements'
    
    def get_queryset(self):
        queryset = Announcement.objects.all()
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
        return queryset.order_by('-published_at')

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        self._log_activity('CREATE', instance, f"Created Announcement: {instance.title}", self.request.user)


class ServiceRequestViewSet(viewsets.ModelViewSet):
    queryset = ServiceRequest.objects.all()
    serializer_class = ServiceRequestSerializer
    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'welfare'
    
    def get_queryset(self):
        queryset = ServiceRequest.objects.all()
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
        return queryset.order_by('-created_at')
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        service_request = self.get_object()
        new_status = request.data.get('status')
        admin_notes = request.data.get('admin_notes', '')
        
        if new_status in dict(ServiceRequest.Status.choices):
            service_request.status = new_status
            service_request.admin_notes = admin_notes
            service_request.handled_by = request.user
            service_request.save()
            return Response(ServiceRequestSerializer(service_request).data)
        
        return Response({'error': 'Invalid status'}, status=400)


# ============================================================================
# PAYMENT API
# PAYMENT API
# PAYMENT API
class PortalPaymentOrderView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import razorpay
        import requests
        import uuid
        
        # Get Tenant Config
        config = MembershipConfig.objects.filter(is_active=True).first()
        if not config:
             return Response({'error': 'Membership config missing'}, status=400)
             
        provider = config.payment_gateway_provider
        
        if provider == MembershipConfig.GatewayProvider.NONE:
             return Response({'error': 'Online payments are currently disabled.'}, status=400)

        amount = request.data.get('amount')
        if not amount or float(amount) < 1:
             return Response({'error': 'Invalid amount'}, status=400)
             
        # 1. RAZORPAY FLOW
        if provider == MembershipConfig.GatewayProvider.RAZORPAY:
            if not config.razorpay_key_id:
                return Response({'error': 'Razorpay not configured'}, status=400)
                
            client = razorpay.Client(auth=(config.razorpay_key_id, config.razorpay_key_secret))
            try:
                payment = client.order.create({'amount': int(float(amount) * 100), 'currency': config.currency, 'payment_capture': '1'})
                return Response({
                    'provider': 'RAZORPAY',
                    'order_id': payment['id'],
                    'amount': payment['amount'],
                    'currency': payment['currency'],
                    'key_id': config.razorpay_key_id
                })
            except Exception as e:
                return Response({'error': str(e)}, status=500)

        # 2. CASHFREE FLOW
        elif provider == MembershipConfig.GatewayProvider.CASHFREE:
            if not config.cashfree_app_id:
                return Response({'error': 'Cashfree not configured'}, status=400)
            
            # Use Sandbox for Development (User requested test creds context)
            # Ideally use a toggle. Defaulting to SANDBOX. 
            base_url = "https://sandbox.cashfree.com/pg"
            
            headers = {
                "x-client-id": config.cashfree_app_id,
                "x-client-secret": config.cashfree_secret_key,
                "x-api-version": "2023-08-01",
                "Content-Type": "application/json"
            }
            
            # Get User Info
            username = request.user.username
            phone = "9999999999" # Default
            if username.startswith('member_'):
                 try:
                     hid = int(username.split('_')[1])
                     h = Household.objects.get(id=hid)
                     if h.phone_number:
                         phone = h.phone_number
                 except:
                     pass
            
            customer_id = f"cust_{username}"
            order_id = f"order_{uuid.uuid4().hex[:10]}"
            
            # Extract PAN from request to embed in Return URL (for persistence across redirect)
            donor_pan = request.data.get('donor_pan', '')
            
            payload = {
                "order_amount": float(amount),
                "order_currency": config.currency,
                "customer_details": {
                    "customer_id": customer_id,
                    "customer_phone": phone,
                    "customer_name": request.user.first_name or username
                },
                "order_meta": {
                    "return_url": f"{request.scheme}://{request.get_host()}/portal/dashboard?order_id={order_id}&pan={donor_pan}"
                }
            }
            
            try:
                resp = requests.post(f"{base_url}/orders", json=payload, headers=headers)
                if resp.status_code == 200:
                    data = resp.json()
                    return Response({
                        'provider': 'CASHFREE',
                        'payment_session_id': data['payment_session_id'],
                        'order_id': data['order_id'],
                        'env': 'SANDBOX'
                    })
                else:
                    return Response({'error': f"Cashfree Error: {resp.text}"}, status=400)
            except Exception as e:
                return Response({'error': str(e)}, status=500)
                
        return Response({'error': 'Invalid Provider'}, status=400)


class PortalPaymentVerifyView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        import razorpay
        import requests
        
        config = MembershipConfig.objects.filter(is_active=True).first()
        if not config:
             return Response({'error': 'Config missing'}, status=400)
             
        data = request.data
        provider = config.payment_gateway_provider
        
        # Get Household
        username = request.user.username
        household = None
        if username.startswith('member_'):
             try:
                 household_id = int(username.split('_')[1])
                 household = Household.objects.get(id=household_id)
             except:
                 pass
        
        if not household:
            return Response({'error': 'Invalid member session'}, status=400)

        # 1. RAZORPAY VERIFY
        if provider == MembershipConfig.GatewayProvider.RAZORPAY:
            try:
                 client = razorpay.Client(auth=(config.razorpay_key_id, config.razorpay_key_secret))
                 params_dict = {
                     'razorpay_order_id': data.get('razorpay_order_id'),
                     'razorpay_payment_id': data.get('razorpay_payment_id'),
                     'razorpay_signature': data.get('razorpay_signature')
                 }
                 client.utility.verify_payment_signature(params_dict)
                 
                 amount = Decimal(str(data.get('amount')))
                 donor_pan = data.get('donor_pan')
                 receipt = MembershipService.process_payment(
                     household, 
                     amount, 
                     notes=f"Razorpay: {data.get('razorpay_payment_id')}",
                     donor_pan=donor_pan
                 )
                 return Response({'status': 'success', 'receipt': receipt.receipt_number})
            except Exception as e:
                 return Response({'error': str(e)}, status=400)
                 
        # 2. CASHFREE VERIFY
        elif provider == MembershipConfig.GatewayProvider.CASHFREE:
            order_id = data.get('order_id')
            if not order_id:
                return Response({'error': 'Order ID missing'}, status=400)
                
            base_url = "https://sandbox.cashfree.com/pg"
            headers = {
                "x-client-id": config.cashfree_app_id,
                "x-client-secret": config.cashfree_secret_key,
                "x-api-version": "2023-08-01"
            }
            
            try:
                resp = requests.get(f"{base_url}/orders/{order_id}", headers=headers)
                if resp.status_code == 200:
                    order_data = resp.json()
                    if order_data.get('order_status') == 'PAID':
                        amount = Decimal(str(order_data.get('order_amount')))
                        donor_pan = data.get('donor_pan')
                        receipt = MembershipService.process_payment(
                             household, 
                             amount, 
                             notes=f"Cashfree: {order_id}",
                             donor_pan=donor_pan
                        )
                        return Response({'status': 'success', 'receipt': receipt.receipt_number})
                    else:
                        return Response({'error': f"Payment status: {order_data.get('order_status')}"}, status=400)
                else:
                    return Response({'error': "Failed to verify with Cashfree"}, status=400)
            except Exception as e:
                return Response({'error': str(e)}, status=500)

        return Response({'error': 'Provider mismatch'}, status=400)

# USER PROFILE API
# ============================================================================

class UserProfileView(APIView):
    """Get and update user profile."""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        return Response({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'date_joined': user.date_joined,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
        })
    
    def put(self, request):
        user = request.user
        user.first_name = request.data.get('first_name', user.first_name)
        user.last_name = request.data.get('last_name', user.last_name)
        user.save()
        return Response({
            'message': 'Profile updated successfully',
            'first_name': user.first_name,
            'last_name': user.last_name,
        })


class ChangeEmailView(APIView):
    """Change user email with password verification."""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        user = request.user
        new_email = request.data.get('new_email')
        password = request.data.get('password')
        
        if not new_email or not password:
            return Response({'error': 'Email and password are required'}, status=400)
        
        if not user.check_password(password):
            return Response({'error': 'Invalid password'}, status=400)
        
        # Check if email already exists
        from django.contrib.auth import get_user_model
        User = get_user_model()
        if User.objects.filter(email=new_email).exclude(id=user.id).exists():
            return Response({'error': 'Email already in use'}, status=400)
        
        user.email = new_email
        user.save()
        return Response({'message': 'Email updated successfully', 'email': new_email})


class ChangePasswordView(APIView):
    """Change user password."""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        user = request.user
        current_password = request.data.get('current_password')
        new_password = request.data.get('new_password')
        confirm_password = request.data.get('confirm_password')
        
        if not all([current_password, new_password, confirm_password]):
            return Response({'error': 'All password fields are required'}, status=400)
        
        if not user.check_password(current_password):
            return Response({'error': 'Current password is incorrect'}, status=400)
        
        if new_password != confirm_password:
            return Response({'error': 'New passwords do not match'}, status=400)
        
        if len(new_password) < 8:
            return Response({'error': 'Password must be at least 8 characters'}, status=400)
        
        user.set_password(new_password)
        user.save()
        return Response({'message': 'Password changed successfully'})


# ============================================================================
# TELEGRAM NOTIFICATION APIs
# ============================================================================

class TelegramBroadcastAnnouncementView(APIView):
    """Broadcast an announcement to all Telegram-linked members."""
    permission_classes = [IsAdminUser]
    
    def post(self, request):
        from apps.shared.telegram import broadcast_announcement
        
        title = request.data.get('title')
        content = request.data.get('content')
        
        if not title or not content:
            return Response({'error': 'Title and content are required'}, status=400)
        
        result = broadcast_announcement(title, content)
        return Response({
            'message': f"Announcement sent to {result['sent']} members",
            'sent': result['sent'],
            'failed': result['failed']
        })


class TelegramPaymentRemindersView(APIView):
    """Send payment reminders to all pending households with linked Telegram."""
    permission_classes = [IsAdminUser]
    
    def post(self, request):
        from apps.shared.telegram import send_bulk_payment_reminders
        
        portal_url = request.data.get('portal_url')
        result = send_bulk_payment_reminders(portal_url)
        
        return Response({
            'message': f"Reminders sent to {result['sent']} households",
            'sent': result['sent'],
            'failed': result['failed'],
            'skipped': result['skipped']
        })


class TelegramStatsView(APIView):
    """Get Telegram linking stats for the tenant."""
    permission_classes = [IsAdminUser]
    
    def get(self, request):
        from apps.jamath.models import TelegramLink, Household
        
        total_households = Household.objects.count()
        linked_count = TelegramLink.objects.filter(is_verified=True).count()
        pending_renewals = Household.objects.filter(is_membership_active=False).count()
        
        return Response({
            'total_households': total_households,
            'telegram_linked': linked_count,
            'pending_renewals': pending_renewals,
            'link_percentage': round((linked_count / total_households * 100) if total_households else 0, 1)
        })


class TelegramIndividualReminderView(APIView):
    """Send payment reminder to a specific household."""
    permission_classes = [IsAdminUser]
    
    def post(self, request, household_id):
        from apps.shared.telegram import send_individual_reminder
        
        try:
            household = Household.objects.get(id=household_id)
        except Household.DoesNotExist:
            return Response({'error': 'Household not found'}, status=404)

        # 1. Create Announcement (Persist reminder in Portal)
        try:
            head = household.members.filter(is_head_of_family=True).first()
            head_name = head.full_name if head else "Member"
            
            Announcement.objects.create(
                title="Membership Renewal Reminder",
                content=f"Assalamu Alaikum {head_name},\n\nThis is a gentle reminder to renew your membership contribution.\n\nJazakallah Khair,\nDigitalJamath Team",
                status='PUBLISHED',
                target_household=household,
                created_by=request.user if request.user.is_authenticated else None
            )
        except Exception as e:
            # If announcement fails, looking at the error might be useful, but let's proceed to telegram
            print(f"Error creating announcement: {e}")

        # 2. Send Telegram (Notification)
        portal_url = request.data.get('portal_url')
        result = send_individual_reminder(household_id, portal_url)
        
        # We consider it a success if we created the announcement, OR if telegram worked.
        # Since we just created the announcement above, we can return success.
        # But let's check result for debugging context.
        
        message = 'Reminder sent successfully'
        if not result['success']:
             message += f" (Note: Telegram notification failed: {result.get('error')})"
        
        return Response({'message': message})


# ============================================================================
# MIZAN LEDGER VIEWSETS
# ============================================================================

class LedgerViewSet(viewsets.ModelViewSet):
    """Chart of Accounts management."""
    queryset = Ledger.objects.filter(parent=None, is_active=True)  # Only top-level by default
    serializer_class = LedgerSerializer
    
    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'finance'

    def get_queryset(self):
        queryset = Ledger.objects.filter(is_active=True)
        account_type = self.request.query_params.get('type')
        flat = self.request.query_params.get('flat')
        
        if account_type:
            queryset = queryset.filter(account_type=account_type)
        
        # Flat list for dropdowns
        if flat:
            return queryset.order_by('code')
        
        # Hierarchical (top-level only, children via serializer)
        return queryset.filter(parent=None).order_by('code')

    def get_object(self):
        """Override to allow finding any active ledger for detail operations (not just top-level)."""
        queryset = Ledger.objects.filter(is_active=True)
        lookup_url_kwarg = self.lookup_url_kwarg or self.lookup_field
        filter_kwargs = {self.lookup_field: self.kwargs[lookup_url_kwarg]}
        obj = get_object_or_404(queryset, **filter_kwargs)
        self.check_object_permissions(self.request, obj)
        return obj

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.is_system:
            return Response({'error': 'Cannot delete system accounts.'}, status=400)
        if instance.journal_items.exists():
            return Response({'error': 'Cannot delete account with transactions.'}, status=400)
        instance.is_active = False
        instance.save()
        return Response(status=204)


class SupplierViewSet(viewsets.ModelViewSet):
    """Vendor/Supplier management."""
    queryset = Supplier.objects.filter(is_active=True)
    serializer_class = SupplierSerializer
    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'finance'
    


    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.payments.exists():
            return Response({'error': 'Cannot delete supplier with payments.'}, status=400)
        instance.is_active = False
        instance.save()
        return Response(status=204)


class JournalEntryViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """Journal Entry (Voucher) management."""
    queryset = JournalEntry.objects.all().order_by('-date')
    serializer_class = JournalEntrySerializer
    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'finance'

    def perform_create(self, serializer):
        entry = serializer.save(created_by=self.request.user)
        
        action_desc = "Created Journal Entry"
        amount_str = f"₹{entry.total_amount:g}"  # Remove trailing zeros if integer
        
        if entry.voucher_type == 'RECEIPT':
             action_desc = f"Received Payment of {amount_str} ({entry.narration})"
        elif entry.voucher_type == 'PAYMENT':
             action_desc = f"Made Payment of {amount_str} ({entry.narration})"
             
        self._log_activity('CREATE', entry, action_desc, self.request.user)

    def get_queryset(self):
        queryset = JournalEntry.objects.all()
        voucher_type = self.request.query_params.get('type')
        date_from = self.request.query_params.get('from')
        date_to = self.request.query_params.get('to')
        
        if voucher_type:
            queryset = queryset.filter(voucher_type=voucher_type)
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
            
        exclude_zakat = self.request.query_params.get('exclude_zakat')
        if exclude_zakat == 'true':
            queryset = queryset.exclude(items__ledger__fund_type='ZAKAT')
        
        return queryset.order_by('-date', '-created_at')

    def create(self, request, *args, **kwargs):
        """Override create to add payment validation."""
        voucher_type = request.data.get('voucher_type')
        
        # Validate payment amount doesn't exceed available balance
        if voucher_type == 'PAYMENT':
            items_data = request.data.get('items', [])
            
            # Calculate total payment amount (sum of expense debits, excluding Zakat)
            total_payment = Decimal('0')
            is_zakat_payment = False
            
            for item_data in items_data:
                ledger_id = item_data.get('ledger')
                debit_amt = Decimal(str(item_data.get('debit_amount', 0)))
                
                try:
                    ledger = Ledger.objects.get(id=ledger_id)
                    # Check if this is a Zakat expense
                    if ledger.fund_type == 'ZAKAT':
                        is_zakat_payment = True
                    # Sum up expense amounts (debits to expense accounts)
                    if ledger.account_type == 'EXPENSE' and debit_amt > 0:
                        total_payment += debit_amt
                except Ledger.DoesNotExist:
                    pass
            
            # Only validate balance for General payments (not Zakat)
            if not is_zakat_payment and total_payment > 0:
                # Get available balance (Cash + Bank, excluding Zakat)
                cash_ledger = Ledger.objects.filter(code='1001').first()  # Cash
                bank_ledger = Ledger.objects.filter(code='1002').first()  # Bank
                
                available_balance = Decimal('0')
                if cash_ledger:
                    available_balance += cash_ledger.balance
                if bank_ledger:
                    available_balance += bank_ledger.balance
                
                if total_payment > available_balance:
                    return Response({
                        'error': f'Insufficient funds. Available balance: ₹{available_balance}, Payment amount: ₹{total_payment}'
                    }, status=400)
        
        return super().create(request, *args, **kwargs)




    @action(detail=True, methods=['post'])
    def finalize(self, request, pk=None):
        """Lock a journal entry to prevent further modifications."""
        entry = self.get_object()
        if entry.is_finalized:
            return Response({'error': 'Already finalized.'}, status=400)
        entry.is_finalized = True
        entry.save()
        return Response({'message': 'Entry finalized successfully.'})

    @action(detail=True, methods=['post'])
    def reverse(self, request, pk=None):
        """
        Create a reversal entry for this journal entry.
        Swaps debits and credits to cancel the original transaction.
        """
        from django.db import transaction as db_transaction
        
        original = self.get_object()
        
        if original.is_finalized:
            return Response({'error': 'Cannot reverse a finalized entry. Contact administrator.'}, status=400)
        
        try:
            with db_transaction.atomic():
                # Create reversal entry
                reversal = JournalEntry.objects.create(
                    voucher_type=original.voucher_type,
                    date=timezone.now().date(),
                    narration=f"REVERSAL of {original.voucher_number}: {original.narration}",
                    donor=original.donor,
                    donor_name_manual=original.donor_name_manual,
                    supplier=original.supplier,
                    payment_mode=original.payment_mode,
                    created_by=request.user
                )
                
                # Create reversed items (swap debit/credit)
                for item in original.items.all():
                    JournalItem.objects.create(
                        journal_entry=reversal,
                        ledger=item.ledger,
                        debit_amount=item.credit_amount,  # Swap
                        credit_amount=item.debit_amount,  # Swap
                        particulars=f"Reversal: {item.particulars or ''}"
                    )
                
                # Mark original as finalized
                original.is_finalized = True
                original.save()
                
                return Response({
                    'message': 'Entry reversed successfully.',
                    'reversal_voucher': reversal.voucher_number,
                    'reversal_id': reversal.id
                })
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class LedgerReportsView(APIView):
    """Ledger reports: Day Book, Trial Balance."""
    permission_classes = [IsAdminUser]

    def get(self, request, report_type):
        from django.db.models import Sum, Q

        if report_type == 'day-book':
            # Default to Date view (how it was)
            filter_mode = request.query_params.get('mode', 'date') # 'date' or 'month'
            target_date_str = request.query_params.get('date', timezone.now().date().isoformat())
            target_date = timezone.datetime.fromisoformat(target_date_str).date()
            fund_type = request.query_params.get('fund_type')
            
            # Base query
            entries = JournalEntry.objects.exclude(voucher_type='JOURNAL')
            
            date_label = target_date_str
            if filter_mode == 'date':
                entries = entries.filter(date=target_date)
            else:
                # Month mode: Filter from start of that month to end of that month
                start_date = target_date.replace(day=1)
                # Logic to get end of month
                if start_date.month == 12:
                    end_date = start_date.replace(year=start_date.year + 1, month=1, day=1) - timezone.timedelta(days=1)
                else:
                    end_date = start_date.replace(month=start_date.month + 1, day=1) - timezone.timedelta(days=1)
                
                entries = entries.filter(date__gte=start_date, date__lte=end_date)
                date_label = start_date.strftime("%B %Y")

            if fund_type == 'ZAKAT':
                # Strictly Zakat-related entries
                entries = entries.filter(items__ledger__fund_type='ZAKAT').distinct()
                sum_filter = Q(items__ledger__fund_type='ZAKAT')
                
            elif fund_type == 'GENERAL':
                # "General" View: Strictly exclude any transaction that involves Zakat funds.
                # This ensures mixed receipts or pure Zakat receipts are hidden.
                # Transactions like "Activation" (Bank + General Income) will remain visible because they have NO Zakat items.
                entries = entries.exclude(items__ledger__fund_type='ZAKAT')
                
                # Aggregation: Sum everything in the remaining entries
                # (Since we excluded Zakat entries entirely, we don't need a complex sum filter, 
                # but to be safe and consistent with previous logic, we can keep it simple)
                sum_filter = ~Q(items__ledger__fund_type='ZAKAT')
            
            else:
                # ALL
                sum_filter = Q()

            sort_order = request.query_params.get('sort', 'newest')
            if sort_order == 'oldest':
                entries = entries.order_by('date', 'created_at')
            else:
                entries = entries.order_by('-date', '-created_at')
            return Response({
                'date': date_label,
                'entries': JournalEntrySerializer(entries, many=True).data,
                'summary': {
                    'total_receipts': entries.filter(voucher_type='RECEIPT').aggregate(
                        total=Sum('items__credit_amount', filter=sum_filter))['total'] or 0,
                    'total_payments': entries.filter(voucher_type='PAYMENT').aggregate(
                        total=Sum('items__debit_amount', filter=sum_filter))['total'] or 0,
                }
            })

        elif report_type == 'dashboard-stats':
            try:
                today = timezone.now().date()
                start_of_month = today.replace(day=1)
                
                # Income: Sum of CREDITS to INCOME accounts (month only)
                # Note: We include Journal Entries now to catch adjustments
                # Income: Net Credit (Credit - Debit) to INCOME accounts
                income_stats = JournalItem.objects.filter(
                    journal_entry__date__gte=start_of_month,
                    journal_entry__date__lte=today,
                    ledger__account_type='INCOME'
                ).exclude(
                    ledger__fund_type='ZAKAT'
                ).aggregate(
                    credits=Sum('credit_amount'),
                    debits=Sum('debit_amount')
                )
                income_this_month = (income_stats['credits'] or Decimal('0')) - (income_stats['debits'] or Decimal('0'))

                # Expense: Net Debit (Debit - Credit) to EXPENSE accounts
                expense_stats = JournalItem.objects.filter(
                    journal_entry__date__gte=start_of_month,
                    journal_entry__date__lte=today,
                    ledger__account_type='EXPENSE'
                ).exclude(
                    ledger__fund_type='ZAKAT'
                ).aggregate(
                    debits=Sum('debit_amount'),
                    credits=Sum('credit_amount')
                )
                expense_this_month = (expense_stats['debits'] or Decimal('0')) - (expense_stats['credits'] or Decimal('0'))

                # Handling Negative Balances (e.g. Reversals exceeding actual transactions)
                # If Expense is negative (Net Credit), treat it as "Other Income" visually
                if expense_this_month < 0:
                     income_this_month += abs(expense_this_month)
                     expense_this_month = Decimal('0')
                
                # If Income is negative (Net Debit), treat it as "Refund Expense" visually
                if income_this_month < 0:
                     expense_this_month += abs(income_this_month)
                     income_this_month = Decimal('0')

                # Total Available Balance (Cash + Bank, excluding Zakat)
                # 1. Calculate Gross Liquid Assets (All Cash + Bank)
                liquid_assets = JournalItem.objects.filter(
                    ledger__account_type='ASSET',
                    ledger__code__startswith='100'  # Cash & Bank codes
                ).aggregate(
                    debit=Sum('debit_amount'),
                    credit=Sum('credit_amount')
                )
                
                gross_cash = (liquid_assets['debit'] or Decimal('0')) - (liquid_assets['credit'] or Decimal('0'))

                # 2. Calculate Restricted Zakat Balance (Income + Equity - Expense)
                # We MUST include EQUITY to capture Opening Balances/Corpus
                zakat_stats = JournalItem.objects.filter(
                    ledger__fund_type='ZAKAT'
                ).aggregate(
                    income=Sum('credit_amount', filter=Q(ledger__account_type='INCOME')),
                    equity=Sum('credit_amount', filter=Q(ledger__account_type='EQUITY')),
                    expense=Sum('debit_amount', filter=Q(ledger__account_type='EXPENSE'))
                )
                
                z_income = zakat_stats['income'] or Decimal('0')
                z_equity = zakat_stats['equity'] or Decimal('0')
                z_expense = zakat_stats['expense'] or Decimal('0')
                
                zakat_balance = (z_income + z_equity) - z_expense
                
                # General Balance calculation
                # If Zakat Fund is negative (Deficit), it means we have "borrowed" from General Cash to pay Zakat expenses.
                # In that case, the physical cash (gross_cash) has already been reduced by the expense.
                # We should NOT subtract the deficit again (which would increase General Balance mathematically).
                # We should only reserve positive Zakat balances.
                zakat_reserve = max(Decimal('0'), zakat_balance)
                
                general_available = gross_cash - zakat_reserve

                return Response({
                    'income_this_month': str(income_this_month),
                    'expense_this_month': str(expense_this_month),
                    'general_balance': str(general_available),
                    'zakat_balance': str(zakat_balance)
                })
            except Exception as e:
                return Response({
                    'error': str(e),
                    'income_this_month': '0',
                    'expense_this_month': '0',
                    'general_balance': '0',
                    'zakat_balance': '0'
                })

        elif report_type == 'trial-balance':
            ledgers = Ledger.objects.filter(is_active=True).order_by('code')
            data = []
            total_debit = Decimal('0.00')
            total_credit = Decimal('0.00')
            
            for ledger in ledgers:
                balance = ledger.balance
                if balance > 0:
                    if ledger.account_type in ['ASSET', 'EXPENSE']:
                        total_debit += balance
                        data.append({'code': ledger.code, 'name': ledger.name, 'debit': balance, 'credit': 0})
                    else:
                        total_credit += balance
                        data.append({'code': ledger.code, 'name': ledger.name, 'debit': 0, 'credit': balance})
                elif balance < 0:
                    if ledger.account_type in ['ASSET', 'EXPENSE']:
                        total_credit += abs(balance)
                        data.append({'code': ledger.code, 'name': ledger.name, 'debit': 0, 'credit': abs(balance)})
                    else:
                        total_debit += abs(balance)
                        data.append({'code': ledger.code, 'name': ledger.name, 'debit': abs(balance), 'credit': 0})
            
            return Response({
                'ledgers': data,
                'total_debit': total_debit,
                'total_credit': total_credit,
                'is_balanced': total_debit == total_credit
            })



        return Response({'error': 'Invalid report type'}, status=400)


class TallyExportView(APIView):
    """Export financial data to Tally-compatible Excel format."""
    permission_classes = [IsAdminUser]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        import io

        # Get date range from query params (default: current financial year Apr-Mar)
        year = request.query_params.get('year')
        if year:
            year = int(year)
        else:
            today = timezone.now().date()
            year = today.year if today.month >= 4 else today.year - 1
        
        start_date = f"{year}-04-01"
        end_date = f"{year + 1}-03-31"

        # Fetch all journal entries in date range
        entries = JournalEntry.objects.filter(
            date__gte=start_date,
            date__lte=end_date
        ).prefetch_related('items__ledger', 'donor').order_by('date', 'voucher_number')

        # Create workbook
        wb = Workbook()
        
        # ========== Tab 1: All Transactions ==========
        ws1 = wb.active
        ws1.title = "Journal Entries"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "Date", "Voucher_Type", "Voucher_Number", "Ledger_Code", "Ledger_Name",
            "Fund_Category", "Debit", "Credit", "Narration", "Payment_Mode",
            "Donor_Name", "Donor_PAN", "Supplier_Name", "Invoice_No", "Created_By"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        row_num = 2
        for entry in entries:
            for item in entry.items.all():
                # Format date as DD-MM-YYYY
                formatted_date = entry.date.strftime("%d-%m-%Y")
                
                donor_name = ""
                donor_pan = ""
                if entry.donor:
                    donor_name = entry.donor.full_name
                elif entry.donor_name_manual:
                    donor_name = entry.donor_name_manual
                donor_pan = entry.donor_pan or ""
                
                supplier_name = entry.supplier.name if entry.supplier else ""
                
                row_data = [
                    formatted_date,
                    entry.voucher_type,
                    entry.voucher_number,
                    item.ledger.code,
                    item.ledger.name,
                    item.ledger.fund_type or "GENERAL",
                    float(item.debit_amount) if item.debit_amount > 0 else "",
                    float(item.credit_amount) if item.credit_amount > 0 else "",
                    entry.narration,
                    entry.payment_mode,
                    donor_name,
                    donor_pan,
                    supplier_name,
                    entry.vendor_invoice_no or "",
                    entry.created_by.username if entry.created_by else ""
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws1.cell(row=row_num, column=col, value=value)
                    cell.border = thin_border
                
                row_num += 1

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws1.column_dimensions[get_column_letter(col)].width = 15

        # ========== Tab 2: Donor List (for Form 10BD) ==========
        ws2 = wb.create_sheet(title="Donor List (Form 10BD)")
        
        donor_headers = [
            "Sr_No", "Donor_Name", "PAN", "Phone", "Total_Donation",
            "Donation_Type", "Address"
        ]
        
        for col, header in enumerate(donor_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Aggregate donors
        donor_data = {}
        for entry in entries.filter(voucher_type='RECEIPT'):
            donor_key = None
            donor_name = ""
            donor_pan = ""
            donor_phone = ""
            
            if entry.donor:
                donor_key = f"member_{entry.donor.id}"
                donor_name = entry.donor.full_name
                donor_pan = entry.donor_pan or ""
                donor_phone = entry.donor.household.phone_number or "" if hasattr(entry.donor, 'household') else ""
            elif entry.donor_name_manual:
                donor_key = f"guest_{entry.donor_name_manual}_{entry.donor_pan or 'nopan'}"
                donor_name = entry.donor_name_manual
                donor_pan = entry.donor_pan or ""
            
            if donor_key:
                if donor_key not in donor_data:
                    donor_data[donor_key] = {
                        'name': donor_name,
                        'pan': donor_pan,
                        'phone': donor_phone,
                        'total': Decimal('0.00'),
                        'types': set()
                    }
                
                # Add total (credit amounts for receipts)
                for item in entry.items.filter(credit_amount__gt=0):
                    donor_data[donor_key]['total'] += item.credit_amount
                    if item.ledger.fund_type:
                        donor_data[donor_key]['types'].add(item.ledger.fund_type)

        # Write donor rows
        row_num = 2
        for idx, (key, data) in enumerate(sorted(donor_data.items(), key=lambda x: x[1]['name']), 1):
            row = [
                idx,
                data['name'],
                data['pan'],
                data['phone'],
                float(data['total']),
                ", ".join(data['types']) or "GENERAL",
                ""
            ]
            for col, value in enumerate(row, 1):
                cell = ws2.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
            row_num += 1

        # Auto-adjust column widths
        for col in range(1, len(donor_headers) + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 18

        # Generate response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Mizan_Export_FY{year}-{year+1}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ============================================================================
# RECEIPT PDF GENERATION
# ============================================================================

class ReceiptPDFView(APIView):
    """Generate PDF receipt for a journal entry (admin) or by receipt ID."""
    permission_classes = [IsAdminUser]
    
    def get(self, request, entry_id):
        from apps.jamath.receipt_generator import generate_receipt_pdf
        
        try:
            entry = JournalEntry.objects.get(id=entry_id)
        except JournalEntry.DoesNotExist:
            return Response({'error': 'Journal entry not found'}, status=404)
        
        # Only receipts can generate PDFs
        if entry.voucher_type != 'RECEIPT':
            return Response({'error': 'Only receipt vouchers can generate PDFs'}, status=400)
        
        # Get organization config
        config = MembershipConfig.objects.filter(is_active=True).first()
        
        # Calculate amount from journal items (credit side)
        amount = sum(item.credit_amount for item in entry.items.all())
        
        # Generate receipt number if not exists
        receipt_number = f"RCP-{entry.date.strftime('%Y%m%d')}-{entry.id:04d}"
        
        # Generate PDF
        pdf_bytes = generate_receipt_pdf(
            receipt_number=receipt_number,
            payment_date=entry.date,
            donor_name=entry.donor_name or entry.narration or "Member",
            donor_address="",
            donor_pan=entry.donor_pan or "",
            amount=amount,
            membership_portion=amount,  # Can be split if needed
            donation_portion=0,
            payment_mode=entry.payment_mode or "Online",
            org_name=config.organization_name if config else "Digital Jamath",
            org_address=config.organization_address if config else "",
            org_pan=config.organization_pan if config else "",
            reg_80g=config.registration_number_80g if config else "",
            masjid_name=config.masjid_name if config else "",
        )
        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Receipt_{receipt_number}.pdf"'
        return response


class PortalReceiptListView(APIView):
    """List receipts for the logged-in household (portal)."""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        if not request.user.username.startswith('member_'):
            return Response({'error': 'Not authenticated as member'}, status=401)
        
        household_id = int(request.user.username.split('_')[1])
        
        try:
            household = Household.objects.get(id=household_id)
        except Household.DoesNotExist:
            return Response({'error': 'Household not found'}, status=404)
        
        # Find receipt entries for this household
        entries = JournalEntry.objects.filter(
            voucher_type='RECEIPT',
            donor__household_id=household_id
        ).order_by('-date')
        
        receipts = []
        for entry in entries:
            amount = sum(item.credit_amount for item in entry.items.all())
            receipts.append({
                'id': entry.id,
                'receipt_number': f"RCP-{entry.date.strftime('%Y%m%d')}-{entry.id:04d}",
                'date': entry.date.isoformat(),
                'amount': float(amount),
                'description': entry.narration,
                'payment_mode': entry.payment_mode or 'Online',
            })
        
        return Response(receipts)


class PortalReceiptPDFView(APIView):
    """Download PDF receipt for portal user."""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, entry_id):
        from apps.jamath.receipt_generator import generate_receipt_pdf
        
        if not request.user.username.startswith('member_'):
            return Response({'error': 'Not authenticated as member'}, status=401)
        
        household_id = int(request.user.username.split('_')[1])
        
        try:
            entry = JournalEntry.objects.get(id=entry_id, donor__household_id=household_id)
        except JournalEntry.DoesNotExist:
            return Response({'error': 'Receipt not found'}, status=404)
        
        if entry.voucher_type != 'RECEIPT':
            return Response({'error': 'Invalid receipt'}, status=400)
        
        # Get organization config
        config = MembershipConfig.objects.filter(is_active=True).first()
        household = entry.household
        head = household.members.filter(is_head_of_family=True).first() if household else None
        
        amount = sum(item.credit_amount for item in entry.items.all())
        receipt_number = f"RCP-{entry.date.strftime('%Y%m%d')}-{entry.id:04d}"
        
        pdf_bytes = generate_receipt_pdf(
            receipt_number=receipt_number,
            payment_date=entry.date,
            donor_name=head.full_name if head else entry.donor_name or "Member",
            donor_address=household.address if household else "",
            donor_pan=entry.donor_pan or "",
            amount=amount,
            membership_portion=amount,
            donation_portion=0,
            payment_mode=entry.payment_mode or "Online",
            org_name=config.organization_name if config else "Digital Jamath",
            org_address=config.organization_address if config else "",
            org_pan=config.organization_pan if config else "",
            reg_80g=config.registration_number_80g if config else "",
            masjid_name=config.masjid_name if config else "",
        )
        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Receipt_{receipt_number}.pdf"'
        return response
class ReminderViewSet(viewsets.ViewSet):
    """ViewSet to manage reminders and custom messages via portal announcements."""
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'])
    def send(self, request):
        """Send a reminder or custom message as a targeted announcement."""
        household_id = request.data.get('household_id')
        message_type = request.data.get('type', 'REMINDER') # REMINDER or MESSAGE
        custom_text = request.data.get('message', '')

        try:
            household = Household.objects.get(id=household_id)
        except Household.DoesNotExist:
            return Response({'error': 'Household not found'}, status=404)

        head = household.members.filter(is_head_of_family=True).first()
        head_name = head.full_name if head else "Member"

        if message_type == 'REMINDER':
            title = "Membership Renewal Reminder"
            content = f"""Assalamu Alaikum {head_name},

Kindly renew your membership.

Jazakallah Khair
— DigitalJamath Office"""

        elif message_type == 'MESSAGE':
            if not custom_text:
                 return Response({'error': 'Message text is required'}, status=400)

            title = "Message from Jamath Office"
            content = f"""Assalamu Alaikum {head_name},

{custom_text}

— DigitalJamath Office"""

        else:
            return Response({'error': 'Invalid message type'}, status=400)

        # Create targeted announcement
        try:
            announcement = Announcement.objects.create(
                title=title,
                content=content,
                status='PUBLISHED',
                target_household=household,
                created_by=request.user
            )
            
            # FUTURE: Integration point for Telegram Notifications
            # if household.telegram_chat_id:
            #     send_telegram_message(household.telegram_chat_id, content)

            return Response({
                'success': True,
                'message': f'Announcement sent successfully to {head_name}',
                'announcement_id': announcement.id
            })
        except Exception as e:
            return Response({'error': f'Failed to create announcement: {str(e)}'}, status=500)

# ============================================================================
# ADMIN ACTIVITY LOGS
# ============================================================================

class ActivityLogStaffSourceView(APIView):
    """
    Returns a list of users who are relevant for Activity Log filtering:
    1. Superusers (Admin)
    2. Active Staff
    3. Users who have at least one ActivityLog entry (Historical)
    """
    permission_classes = [IsAdminUser | HasStaffPermission]

    def get(self, request):
        # 1. Get all relevant user IDs
        admin_ids = User.objects.filter(is_superuser=True).values_list('id', flat=True)
        staff_ids = StaffMember.objects.values_list('user_id', flat=True)
        log_user_ids = ActivityLog.objects.exclude(user__isnull=True).values_list('user_id', flat=True).distinct()
        
        # Combine distinct IDs
        all_ids = set(admin_ids) | set(staff_ids) | set(log_user_ids)
        
        # Fetch Users
        users = User.objects.filter(id__in=all_ids).select_related('staff_profile', 'staff_profile__role')
        
        results = []
        for user in users:
            # Exclude debug_admin explicitly
            if user.username == 'debug_admin':
                continue

            # Determine type
            role_type = "User"
            if user.is_superuser:
                role_type = "Admin"
            elif hasattr(user, 'staff_profile'):
                role_type = "Staff"
                if user.staff_profile.role:
                     role_type = f"Staff ({user.staff_profile.role.name})"
            
            # Filter out generic "User" type (neither Admin nor Staff)
            if role_type == "User":
                continue

            # Format Name: Full Name > Username
            name = user.get_full_name() or user.username
            
            results.append({
                'id': user.id,
                'name': name,
                'username': user.username,
                'type': role_type,
                'is_active': user.is_active
            })
            
        # Sort: Admin first, then by name
        results.sort(key=lambda x: (0 if x['type'] == 'Admin' else 1, x['name']))
        
        return Response(results)

class ActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    """Admin-only view of staff activity logs."""
    # Filter to show only Staff (either Django admin or App staff)
    queryset = ActivityLog.objects.filter(
        models.Q(user__is_superuser=True) | 
        models.Q(user__staff_profile__isnull=False)
    ).distinct().order_by('-timestamp', '-id')

    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'settings' # Or users/jamath?
    
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['details', 'user__username', 'model_name', 'action']
    ordering_fields = ['timestamp', 'action']

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # 1. Action Filter
        action = self.request.query_params.get('action')
        if action and action != 'ALL':
            queryset = queryset.filter(action=action)

        # 2. Module Filter
        module = self.request.query_params.get('module')
        if module and module != 'ALL':
            queryset = queryset.filter(module=module)

        # 3. Staff Filter (by ID)
        staff_id = self.request.query_params.get('staff_id')
        if staff_id and staff_id != 'ALL':
            queryset = queryset.filter(user_id=staff_id)

        # 4. Date Range Filter
        start_date = self.request.query_params.get('start_date')
        if start_date:
            from django.utils.dateparse import parse_date
            s_date = parse_date(start_date)
            if s_date:
                queryset = queryset.filter(timestamp__date__gte=s_date)
        
        end_date = self.request.query_params.get('end_date')
        if end_date:
            from django.utils.dateparse import parse_date
            e_date = parse_date(end_date)
            if e_date:
                queryset = queryset.filter(timestamp__date__lte=e_date)

        return queryset

    permission_classes = [IsAdminUser | HasStaffPermission]
    required_module = 'settings' # Or users/jamath?
    
    class ActivityLogSerializer(serializers.ModelSerializer):
        username = serializers.CharField(source='user.username', read_only=True)
        class Meta:
            model = ActivityLog
            fields = '__all__'

    serializer_class = ActivityLogSerializer

class DashboardStatsView(APIView):
    """Aggregated stats for the dashboard."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from django.utils import timezone
            from django.db import models
            from decimal import Decimal
            from apps.jamath.models import Household, Member, JournalItem
            
            now = timezone.now()
            month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            
            # Households
            total_households = Household.objects.count()
            total_members = Member.objects.count()
            # Fix: Filter via related subscriptions instead of property
            pending_renewals = Household.objects.exclude(
                subscriptions__status='ACTIVE',
                subscriptions__end_date__gte=now.date()
            ).count()
            
            # Finance Aggregates (Robust P&L Logic)
            # Income: Net Credit (Credit - Debit) to INCOME accounts
            income_stats = JournalItem.objects.filter(
                journal_entry__date__gte=month_start.date(),
                journal_entry__date__lte=now.date(),
                ledger__account_type='INCOME'
            ).exclude(
                ledger__fund_type='ZAKAT'
            ).aggregate(
                credits=models.Sum('credit_amount'),
                debits=models.Sum('debit_amount')
            )
            income_this_month = (income_stats['credits'] or Decimal('0')) - (income_stats['debits'] or Decimal('0'))

            # Expense: Net Debit (Debit - Credit) to EXPENSE accounts
            expense_stats = JournalItem.objects.filter(
                journal_entry__date__gte=month_start.date(),
                journal_entry__date__lte=now.date(),
                ledger__account_type='EXPENSE'
            ).exclude(
                ledger__fund_type='ZAKAT'
            ).aggregate(
                debits=models.Sum('debit_amount'),
                credits=models.Sum('credit_amount')
            )
            expense_this_month = (expense_stats['debits'] or Decimal('0')) - (expense_stats['credits'] or Decimal('0'))

            # Handle Reversals/Negative Balances
            if expense_this_month < 0:
                 income_this_month += abs(expense_this_month)
                 expense_this_month = Decimal('0')
            
            if income_this_month < 0:
                 expense_this_month += abs(income_this_month)
                 income_this_month = Decimal('0')

            # Total Lifetime Income (For stats) - Keep simple or update? 
            # Let's keep it simple (Total Receipts) for now as "Total Income" usually implies Revenue+Capital in simple terms
            # Or use pure Income
            total_income_qs = JournalItem.objects.filter(ledger__account_type='INCOME').aggregate(
                c=models.Sum('credit_amount'), d=models.Sum('debit_amount')
            )
            total_income = (total_income_qs['c'] or 0) - (total_income_qs['d'] or 0)

            return Response({
                'households': total_households,
                'members': total_members,
                'pending_renewals': pending_renewals,
                'income_this_month': income_this_month,
                'expense_this_month': expense_this_month,
                'total_income': total_income
            })
        except Exception as e:
            # Return error message in fields to help debug if it happens again
            return Response({
                'households': 0, 'members': 0, 'pending_renewals': 0,
                'income_this_month': 0, 'expense_this_month': 0, 'total_income': 0,
                'debug_error': str(e)
            })
